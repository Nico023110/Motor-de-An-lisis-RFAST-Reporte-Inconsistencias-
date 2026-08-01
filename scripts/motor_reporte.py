# -*- coding: utf-8 -*-

import pandas as pd
import os
import glob
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

# =============================================================================
# INICIO
# =============================================================================

print("\n" + "=" * 70)
print("INICIO REPORTE DE INCONSISTENCIAS")
print("=" * 70)

# =============================================================================
# CARGA DE ARCHIVO
# =============================================================================

ruta_fuente = r"C:\proyecto\reporte_inconsistencias\fuente"
archivos_excel = glob.glob(os.path.join(ruta_fuente, "*.xlsx"))

if not archivos_excel:
    raise FileNotFoundError(
        f"No se encontraron archivos Excel en la carpeta:\n{ruta_fuente}"
    )

# Obtener el archivo más reciente por fecha de modificación
ruta_archivo = max(archivos_excel, key=os.path.getmtime)

print(f"\nArchivo más reciente seleccionado: {os.path.basename(ruta_archivo)}")
print("Cargando archivo...")

dt = pd.read_excel(
    ruta_archivo,
    dtype=str
)

dt.columns = [str(c).strip().lower() for c in dt.columns]

# Soportar archivos con cabeceras truncadas (estilo DBF)
renombres = {
    "finalidad_": "finalidad_rips",
    "nombre_fin": "nombre_finalidad",
    "nombre_cex": "nombre_cexterna",
    "nombre_caj": "nombre_cajero",
    "nombre_pre": "nombre_prestador",
    "nombre_cen": "nombre_centroproduccion",
    "profesiona": "profesional",
    "nombre_pro": "nombre_profesional",
    "dx_princip": "dx_principal",
    "nombre_dx_": "nombre_dx_principal",
    "dx_relacio": "dx_relacionado",
    "nombre_dx2": "nombre_dx_relacionado"
}
dt.rename(columns=renombres, inplace=True)

dt = dt.fillna("")

# Corrección de codificación y consistencia de nombres de IPS
if "nombre_prestador" in dt.columns:
    dt["nombre_prestador"] = (
        dt["nombre_prestador"]
        .astype(str)
        .str.replace("\ufffd", "Ñ", regex=False)
        .str.replace("Ð", "Ñ", regex=False)
        .str.replace(" VISTA HERMOSA", " VISTAHERMOSA", regex=False)
    )

print("Archivo cargado correctamente")
print(f"Registros encontrados: {len(dt):,}")
print(f"Columnas encontradas: {len(dt.columns)}")

# =============================================================================
# ESTRUCTURA DEL REPORTE DE INCONSISTENCIAS
# =============================================================================

columnas_origen = [
    "codigo",
    "nombre",
    "finalidad",
    "finalidad_rips",
    "nombre_finalidad",
    "cexterna",
    "nombre_cexterna",
    "atencion",
    "cajero",
    "nombre_cajero",
    "nombre_prestador",
    "centroprod",
    "nombre_centroproduccion",
    "profesional",
    "nombre_profesional",
    "documento",
    "factura",
    "dx_principal",
    "nombre_dx_principal",
    "dx_relacionado",
    "nombre_dx_relacionado"
]

columnas_reporte = [
    "CODIGO",
    "NOMBRE",
    "FINALIDAD",
    "FINALIDAD_RIPS",
    "NOMBRE_FINALIDAD",
    "CEXTERNA",
    "NOMBRE_CEX",
    "ATENCION",
    "CAJERO",
    "NOMBRE_CAJ",
    "NOMBRE_PRE",
    "CENTROPROD",
    "NOMBRE_CEN",
    "PROFESIONA",
    "NOMBRE_PRO",
    "DOCUMENTO",
    "FACTURA",
    "DX_PRINCIP",
    "NOMBRE_DX_",
    "DX_RELACIO",
    "NOMBRE_DX2",
    "Inconsistencia a Corregir"
]

# =============================================================================
# DATAFRAME DE INCONSISTENCIAS (INICIA VACÍO)
# =============================================================================

dt_inconsistencias = pd.DataFrame(columns=columnas_reporte)

# =============================================================================
# DICCIONARIO DE ZONAS
# =============================================================================

zonas = {

    "ZonaComuna-01": [
        "C.S. TERRON COLORADO",
        "P.S. BELLAVISTA",
        "P.S. VISTAHERMOSA",
        "P.S. LA PAZ"
    ],

    "ZonaComuna-03": [
        "P.S. FRAY DAMIAN",
        "HOSPITAL CAÑAVERALEJO"
    ],

    "ZonaComuna-17": [
        "C.S. PRIMERO DE MAYO"
    ],

    "ZonaComuna-18": [
        "C.S. MELENDEZ",
        "P.S. ALTO POLVORINES",
        "P.S. ALTO NAPOLES",
        "P.S. NAPOLES",
        "P.S. POLVORINES",
        "P.S. LOURDES"
    ],

    "ZonaComuna-20": [
        "C.S. SILOE",
        "P.S. BELEN",
        "P.S. BRISAS DE MAYO",
        "P.S. LA ESTRELLA",
        "P.S. LA SIRENA",
        "P.S. LA SULTANA"
    ],

    "ZonaRuralNorte": [
        "P.S. MONTEBELLO",
        "P.S. EL SALADITO",
        "P.S. LA ELVIRA",
        "P.S. FELIDIA",
        "P.S. PENAS BLANCAS",
        "P.S. PICHINDE",
        "P.S. GOLONDRINAS",
        "P.S. LA LEONERA",
        "P.S. LA PAZ RURAL",
        "P.S. ALTO AGUACATAL",
        "P.S. LA CASTILLA",
        "P.S. LOS ANDES"
    ],

    "ZonaRuralSur": [
        "P.S. LA BUITRERA",
        "P.S. VILLACARMELO",
        "P.S. PANCE",
        "P.S. LA VORAGINE",
        "P.S. EL HORMIGUERO",
        "P.S. CASCAJAL"
    ]

}

print("\n" + "=" * 70)
print("DATAFRAME DE INCONSISTENCIAS")
print("=" * 70)

print("DataFrame creado correctamente")
print(f"Columnas definidas: {len(columnas_reporte)}")
print(f"Registros iniciales: {len(dt_inconsistencias)}")

# =============================================================================
# FUNCIÓN PARA AGREGAR INCONSISTENCIAS
# =============================================================================

def agregar_inconsistencia(registros, mensaje):

    global dt_inconsistencias

    if registros.empty:
        return

    temp = registros[columnas_origen].copy()

    temp.columns = columnas_reporte[:-1]

    temp["Inconsistencia a Corregir"] = mensaje

    # Concatenar en bloque (mucho más rápido que iterar por filas)
    dt_inconsistencias = pd.concat(
        [dt_inconsistencias, temp],
        ignore_index=True
    )
    
    # Eliminar duplicados manteniendo solo la primera inconsistencia detectada por factura/código
    dt_inconsistencias.drop_duplicates(subset=["FACTURA", "CODIGO"], keep="first", inplace=True)

    print(
        f"Total acumulado en reporte: "
        f"{len(dt_inconsistencias):,}"
    )
# =============================================================================
# VALIDACIÓN 01 - CAUSA EXTERNA INCORRECTA PYM
# =============================================================================

def validar_01_causa_externa_pym():

    print("\n" + "-" * 70)
    print("VALIDACIÓN 01 - CAUSA EXTERNA INCORRECTA PYM")
    print("-" * 70)

    mascara = (

        (dt["cexterna"].astype(str).str.strip() == "38")

        &

        (dt["finalidad_rips"].astype(str).str.strip() == "11")

        &

        (
            dt["nombre_centroproduccion"]
            .astype(str)
            .str.upper()
            .str.contains(
                "CURSO DE VIDA|PLANIFICACION FAMILIAR",
                na=False
            )
        )

        &

        (
            dt["nombre"]
            .astype(str)
            .str.upper()
            .str.contains(
                "PRIMERA VEZ|SEGUIMIENTO",
                na=False
            )
        )

    )

    registros_error = dt[mascara].copy()

    print(f"Registros encontrados: {len(registros_error):,}")

    agregar_inconsistencia(
        registros_error,
        "La causa externa no puede ser la 38 (Enfermedad General), debe ser la 40 (Promoción y mantenimiento)."
    )

# =============================================================================
# VALIDACIÓN 02 - FINALIDAD INCORRECTA EN PROGRAMAS DE CONTROL
# =============================================================================

def validar_02_finalidad_programas_control():

    print("\n" + "-" * 70)
    print("VALIDACIÓN 02 - FINALIDAD INCORRECTA EN PROGRAMAS DE CONTROL")
    print("-" * 70)

    mascara = (

    (
        dt["centroprod"]
        .astype(str)
        .str.upper()
        .str.contains(
            "1415|1416|1417",
            na=False
        )
    )

    &

    (
        dt["nombre"]
        .astype(str)
        .str.upper()
        .str.contains(
            "CONTROL|SEGUIMIENTO",
            na=False
        )
    )

    &

    (
        ~dt["finalidad_rips"]
        .astype(str)
        .str.strip()
        .isin(
            ["16", "17", "23", "0", "28"]
        )
    )

    &

    (
        (dt["dx_relacionado"].fillna("").str.strip() == "")

        |

        (
            dt["dx_relacionado"]
            .fillna("")
            .str.strip()
            ==
            dt["dx_principal"]
            .fillna("")
            .str.strip()
        )
    )

)

    registros_error = dt[mascara].copy()

    print(f"Registros encontrados: {len(registros_error):,}")

    agregar_inconsistencia(
        registros_error,
        "La finalidad debe ser 28 (tratamiento) porque son consultas de control y seguimiento de pacientes con diagnósticos ya definidos (ej. hipertensión)."
    )

# =============================================================================
# VALIDACIÓN 03 - FINALIDAD INCORRECTA EN CONSULTAS DE PRIMERA VEZ
# =============================================================================

def validar_03_finalidad_primera_vez():

    print("\n" + "-" * 70)
    print("VALIDACIÓN 03 - FINALIDAD INCORRECTA EN CONSULTAS DE PRIMERA VEZ")
    print("-" * 70)

    mascara = (

        (
            dt["nombre"]
            .astype(str)
            .str.upper()
            .str.contains(
                "PRIMERA VEZ",
                na=False
            )
        )

        &

        (
            dt["centroprod"]
            .astype(str)
            .str.strip()
            .isin(
                ["1415", "1416", "1417"]
            )
        )

        &

        (
            ~dt["finalidad_rips"]
            .astype(str)
            .str.strip()
            .isin(
                ["15", "23", "0"]
            )
        )

    )

    registros_error = dt[mascara].copy()

    print(f"Registros encontrados: {len(registros_error):,}")

    agregar_inconsistencia(
        registros_error,
        "La finalidad debe ser 27 (diagnóstico) porque es una consulta por primera vez, el objetivo principal es evaluar al paciente para confirmar o definir un diagnóstico."
    )
    
# =============================================================================
# VALIDACIÓN 04 - FINALIDAD INCORRECTA EN CONTROLES DE ODONTOLOGÍA
# =============================================================================

def validar_04_finalidad_control_odontologia():

    print("\n" + "-" * 70)
    print("VALIDACIÓN 04 - FINALIDAD INCORRECTA EN CONTROLES DE ODONTOLOGÍA")
    print("-" * 70)

    mascara = (

        (
            dt["nombre"]
            .astype(str)
            .str.upper()
            .str.contains(
                "CONTROL",
                na=False
            )
        )

        &

        (
            dt["centroprod"]
            .astype(str)
            .str.strip()
            .isin(
                ["1300", "1303"]
            )
        )

        &

        (
            ~dt["finalidad_rips"]
            .astype(str)
            .str.strip()
            .isin(
                ["16", "17", "0", "23"]
            )
        )

    )

    registros_error = dt[mascara].copy()

    print(f"Registros encontrados: {len(registros_error):,}")

    agregar_inconsistencia(
        registros_error,
        "Debe ser finalidad 28 (Tratamiento) porque la consulta de control o seguimiento en odontología hace parte de la continuidad del manejo clínico del paciente."
    )

# =============================================================================
# VALIDACIÓN 05 - FINALIDAD INCORRECTA EN PLANIFICACIÓN FAMILIAR
# =============================================================================

def validar_05_planificacion_familiar():

    print("\n" + "-" * 70)
    print("VALIDACIÓN 05 - FINALIDAD INCORRECTA EN PLANIFICACIÓN FAMILIAR")
    print("-" * 70)

    mascara = (

        (
            dt["nombre"]
            .astype(str)
            .str.upper()
            .str.contains(
                "CONSULTA",
                na=False
            )
        )

        &

        (
            dt["centroprod"]
            .astype(str)
            .str.strip()
            .isin(
                ["1405"]
            )
        )

        &

        (
            ~dt["finalidad_rips"]
            .astype(str)
            .str.strip()
            .isin(
                ["19", "21", "23", "25", "0"]
            )
        )

    )

    registros_error = dt[mascara].copy()

    print(f"Registros encontrados: {len(registros_error):,}")

    agregar_inconsistencia(
        registros_error,
        "Debe ser finalidad 31 (Planificación familiar y anticoncepción) y la causa externa 40 (Promoción y mantenimiento) porque la consulta corresponde a una atención de primera vez dentro del programa de planificación familiar (PYP PF)."
    )
    

# =============================================================================
# VALIDACIÓN 06 - FINALIDAD INCORRECTA EN PROGRAMAS DE DETECCIÓN TEMPRANA
# =============================================================================

def validar_06_deteccion_temprana():

    print("\n" + "-" * 70)
    print("VALIDACIÓN 06 - FINALIDAD INCORRECTA EN PROGRAMAS DE DETECCIÓN TEMPRANA")
    print("-" * 70)

    mascara = (

        (
            dt["nombre"]
            .astype(str)
            .str.upper()
            .str.contains(
                "CONSULTA",
                na=False
            )
        )

        &

        (
            dt["centroprod"]
            .astype(str)
            .str.strip()
            .isin(
                ["1408", "1409", "1439", "1440"]
            )
        )

        &

        (
            ~dt["finalidad_rips"]
            .astype(str)
            .str.strip()
            .isin(
                ["12", "15", "16", "23", "0"]
            )
        )

    )

    registros_error = dt[mascara].copy()

    print(f"Registros encontrados: {len(registros_error):,}")

    agregar_inconsistencia(
        registros_error,
        "La finalidad debe ser 24 (detección temprana de enfermedad general) porque las atenciones corresponden a consultas de primera vez dentro de programas de detección de cáncer, cuyo objetivo es identificar de manera oportuna posibles enfermedades, no realizar tratamiento ni solo valoración general."
    )
    

# =============================================================================
# VALIDACIÓN 07 - FINALIDAD INCORRECTA EN EDUCACIÓN INDIVIDUAL
# =============================================================================

def validar_07_educacion_individual():

    print("\n" + "-" * 70)
    print("VALIDACIÓN 07 - FINALIDAD INCORRECTA EN EDUCACIÓN INDIVIDUAL")
    print("-" * 70)

    mascara = (

        (
            dt["nombre"]
            .astype(str)
            .str.upper()
            .str.contains(
                "EDUCACION INDIVIDUAL",
                na=False
            )
        )

        &

        (
            ~dt["finalidad_rips"]
            .astype(str)
            .str.strip()
            .isin(
                [
                    "0", "19", "20", "23", "28", "29", "30",
                    "32", "33", "34", "38", "39", "40",
                    "41", "42"
                ]
            )
        )

    )

    registros_error = dt[mascara].copy()

    print(f"Registros encontrados: {len(registros_error):,}")

    agregar_inconsistencia(
        registros_error,
        "La finalidad registrada no corresponde a una actividad de educación individual. Estas atenciones deben registrarse con una finalidad de Promoción de la Salud (finalidades 40 a 54), ya que su objetivo es la educación y promoción, no el diagnóstico, tratamiento o seguimiento clínico."
    )
# =============================================================================
# COLUMNAS DEL ARCHIVO
# =============================================================================

print("\nColumnas encontradas:")

for columna in dt.columns:
    print(f" - {columna}")

# =============================================================================
# INFORMACIÓN GENERAL
# =============================================================================

print("\n" + "=" * 70)
print("RESUMEN")
print("=" * 70)

print("DataFrame principal: dt")
print("DataFrame inconsistencias: dt_inconsistencias")

print(f"\nDimensiones dt: {dt.shape}")
print(f"Dimensiones dt_inconsistencias: {dt_inconsistencias.shape}")


# =============================================================================
# CARPETA DE SALIDA
# =============================================================================

ruta_salida = r"C:\proyecto\reporte_inconsistencias\reporte"

os.makedirs(ruta_salida, exist_ok=True)

# =============================================================================
# EJECUCIÓN DE VALIDACIONES
# =============================================================================

validar_01_causa_externa_pym()
validar_02_finalidad_programas_control()
validar_03_finalidad_primera_vez()
validar_04_finalidad_control_odontologia()
validar_05_planificacion_familiar()
validar_06_deteccion_temprana()
validar_07_educacion_individual()

# =============================================================================
# GUARDAR REPORTE
# =============================================================================

fecha = datetime.now().strftime("%Y%m%d_%H%M%S")

nombre_archivo = f"Reporte_Inconsistencias_{fecha}.xlsx"

ruta_reporte = os.path.join(
    ruta_salida,
    nombre_archivo
)

with pd.ExcelWriter(
    ruta_reporte,
    engine="openpyxl"
) as writer:

    # Hoja principal
    dt_inconsistencias.to_excel(
        writer,
        sheet_name="Todas las inconsistencias",
        index=False
    )

    # Una hoja por zona
    for nombre_hoja, ips in zonas.items():

        datos = dt_inconsistencias[
            dt_inconsistencias["NOMBRE_PRE"]
            .str.upper()
            .isin(
                [x.upper() for x in ips]
            )
        ]

        datos.to_excel(
            writer,
            sheet_name=nombre_hoja,
            index=False
        )

# =============================================================================
# FORMATO DEL REPORTE
# =============================================================================

wb = load_workbook(ruta_reporte)

# Color amarillo
relleno_amarillo = PatternFill(
    fill_type="solid",
    start_color="FFFF00",
    end_color="FFFF00"
)

# Fuente roja
fuente_roja = Font(
    color="FF0000"
)

# Recorrer todas las hojas
for ws in wb.worksheets:

    # Encabezados amarillos
    for celda in ws[1]:
        celda.fill = relleno_amarillo

    # Buscar la columna de inconsistencias
    columna_inconsistencia = None

    for celda in ws[1]:

        if celda.value == "Inconsistencia a Corregir":
            columna_inconsistencia = celda.column
            break

    if columna_inconsistencia:

        for fila in range(2, ws.max_row + 1):

            ws.cell(
                row=fila,
                column=columna_inconsistencia
            ).font = fuente_roja

wb.save(ruta_reporte)

print("\n" + "=" * 70)
print("REPORTE GENERADO")
print("=" * 70)
print(f"Registros reportados: {len(dt_inconsistencias):,}")
print(f"Archivo generado en:\n{ruta_reporte}")

print("\nProceso finalizado correctamente.")

print("\nProceso listo para iniciar validaciones.")
